#!/usr/bin/python
# -*- coding: utf-8 -*-

from Integrador import Integrador

import os, shutil, logging
from logging.config import fileConfig
import json
from openpyxl import load_workbook

integrador = Integrador()

wb = load_workbook('./cuentas.xlsx')

sheet = wb.get_sheet_by_name('balance')

for r in range(6, 121):
    cuenta = sheet.cell(row=r, column=10).value
    es_grupo = sheet.cell(row=r, column=9).value
    if cuenta is None: continue
    print("procesando: " + cuenta)
    try:
        saldo = integrador.obtener_saldo_cuenta(cuenta)
    except:
        print("no se encontró la cuenta " + cuenta)
        continue

    if es_grupo == 1:
        sheet.cell(row=r, column=11).value = integrador.obtener_saldo_cuenta(cuenta)

cuentas = integrador.get_cuentas_non_group()


for cuenta in cuentas:
    saldo = integrador.obtener_saldo_cuenta(cuenta["name"])
    saldo = saldo if saldo is not None else 0
    print(cuenta["name"] + ", " + str(saldo))

wb.save('./cuentas.xlsx')

